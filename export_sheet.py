import json
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from logger import logger  # Importing the Logger class from logger.py
logger = logger(identifier='AMZN')
# Path to your JSON file and output Excel file
storage_path = 'storage.json'  # replace with your JSON file path
output_path = 'sheet.xlsx'

# Load JSON data
with open(storage_path, 'r') as file:
    data = json.load(file)

# Initialize workbook and select active sheet
wb = Workbook()
ws = wb.active
ws.title = "Amazon Data"

# Define headers and apply styles
headers = ["Amazon Title", "Amazon Link", "Image", "Product Code", "Price"]
header_fill = PatternFill("solid", fgColor="D3D3D3")
header_font = Font(bold=True)

for col_num, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Set column widths
ws.column_dimensions[get_column_letter(1)].width = 50  # Adjust Amazon Title width
ws.column_dimensions[get_column_letter(2)].width = 15  # Amazon Link
ws.column_dimensions[get_column_letter(3)].width = 15  # Image
ws.column_dimensions[get_column_letter(4)].width = 15  # Product Code
ws.column_dimensions[get_column_letter(5)].width = 15  # Price

# Add data to the sheet
for row_num, (title, link, image_path, code, price) in enumerate(zip(
        data['amazon_titles'],
        data['amazon_links'],
        data['amazon_images'],
        data['product_codes'],
        data['amazon_prices']
), start=2):

    # Skip the entire row if the image is a .webp file
    if image_path.lower().endswith('.webp'):
        continue

    # Amazon Title (left-aligned with wrap text enabled)
    title_cell = ws.cell(row=row_num, column=1, value=title)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)  # Left-aligned with wrap

    # Amazon Link as hyperlink (with black font)
    link_cell = ws.cell(row=row_num, column=2, value="Amazon Link")
    link_cell.hyperlink = link
    link_cell.style = "Hyperlink"
    link_cell.font = Font(color="000000")  # Set link text color to black
    link_cell.alignment = Alignment(horizontal="center", vertical="center")  # Centered both ways

    # Image (skip if it's a .webp file)
    if image_path.lower().endswith('.jpg') and "FMwebp" not in image_path:
        if os.path.exists(image_path):
            try:
                img = Image(image_path)
                img.width, img.height = 50, 50  # Resize image

                # Add image to the worksheet
                ws.add_image(img, f"C{row_num}")

                # Set the row height to match the image height
                ws.row_dimensions[row_num].height = img.height + 10  # Adding extra height for padding

            except Exception as e:
                logger.log("error", f"Error loading image {image_path}")
        else:
            logger.log("error", f"Image file does not exist: {image_path}")

    # Product Code
    code_cell = ws.cell(row=row_num, column=4, value=code)
    code_cell.alignment = Alignment(horizontal="center", vertical="center")  # Center both ways

    # Price
    price_cell = ws.cell(row=row_num, column=5, value=price)
    price_cell.alignment = Alignment(horizontal="center", vertical="center")  # Center both ways

# Save the workbook
try:
    wb.save(output_path)
    logger.log("success", "Data exported to Excel successfully!")
except Exception as e:
    logger.log("error", f"Error saving Excel file: {e}")
